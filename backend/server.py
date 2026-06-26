from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import csv
import logging
import uuid
from pathlib import Path
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
import openpyxl


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


# ----------------------------- Helpers -----------------------------
def parse_num(value) -> Optional[float]:
    """Parse a number that may use comma as decimal separator."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == '':
        return None
    s = s.replace('\xa0', '').replace(' ', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def iva_rate(value) -> float:
    """Normalise an IVA value to a fraction (e.g. 6 -> 0.06, 0.06 -> 0.06)."""
    v = parse_num(value)
    if v is None:
        return 0.0
    return v / 100.0 if v > 1 else v


def round2(x: Optional[float]) -> Optional[float]:
    return round(x, 2) if x is not None else None


# ----------------------------- Cardex (XLS) -----------------------------
def parse_cardex_xlsx(content: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    products = {}
    labs = set()
    rows = ws.iter_rows(min_row=2, values_only=True)
    for r in rows:
        if r is None or len(r) < 18 or r[0] is None:
            continue
        cnp = str(r[0]).strip()
        if cnp == '' or cnp.lower() == 'none':
            continue
        lab = (str(r[5]).strip() if r[5] is not None else 'Sem Laboratório') or 'Sem Laboratório'
        labs.add(lab)
        products[cnp] = {
            'cnp': cnp,
            'descricao': r[1],
            'pva': parse_num(r[2]),
            'pvp': parse_num(r[3]),
            'iva': parse_num(r[4]),
            'laboratorio': lab,
            'acordo_tripartido': parse_num(r[6]),
            'tfo': parse_num(r[8]),
            'simplex': parse_num(r[10]),
            'bolsa': parse_num(r[12]),
            'delegado': parse_num(r[14]),
            'rentb_maxima': parse_num(r[16]),
            'pcu': parse_num(r[17]),
        }
    wb.close()
    return {'products': products, 'labs': sorted(labs)}


@api_router.post("/cardex/upload")
async def upload_cardex(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="O ficheiro deve ser um Excel (.xlsx).")
    content = await file.read()
    try:
        parsed = parse_cardex_xlsx(content)
    except Exception as e:
        logger.exception("cardex parse error")
        raise HTTPException(status_code=400, detail=f"Erro ao processar o Cardex: {e}")

    if not parsed['products']:
        raise HTTPException(status_code=400, detail="Nenhum produto encontrado no ficheiro.")

    doc = {
        '_id': 'current',
        'filename': file.filename,
        'uploaded_at': datetime.now(timezone.utc).isoformat(),
        'count': len(parsed['products']),
        'labs': parsed['labs'],
        'products': parsed['products'],
    }
    await db.cardex.replace_one({'_id': 'current'}, doc, upsert=True)
    return {
        'filename': doc['filename'],
        'uploaded_at': doc['uploaded_at'],
        'count': doc['count'],
        'labs': doc['labs'],
    }


@api_router.get("/cardex")
async def get_cardex():
    doc = await db.cardex.find_one({'_id': 'current'})
    if not doc:
        return {'exists': False}
    sample = list(doc['products'].values())[:50]
    return {
        'exists': True,
        'filename': doc['filename'],
        'uploaded_at': doc['uploaded_at'],
        'count': doc['count'],
        'labs': doc['labs'],
        'sample': sample,
    }


# ----------------------------- Analysis (TXT) -----------------------------
def analyse_txt(content: bytes, cardex_products: dict) -> dict:
    text = content.decode('latin-1')
    reader = csv.reader(io.StringIO(text), delimiter='\t')
    header = next(reader)
    idx = {h.strip(): i for i, h in enumerate(header)}

    required = ['CPR', 'PVP', 'PCU', 'IVA']
    for col in required:
        if col not in idx:
            raise ValueError(f"Coluna obrigatória em falta no TXT: {col}")

    vcols = [idx[f'V({i})'] for i in range(1, 13) if f'V({i})' in idx]

    products = []
    labs_agg = {}
    for row in reader:
        if len(row) <= idx['CPR']:
            continue
        cpr = str(row[idx['CPR']]).strip()
        cx = cardex_products.get(cpr)
        if cx is None:
            continue

        pvp_t = parse_num(row[idx['PVP']])
        pcu_t = parse_num(row[idx['PCU']])
        if pvp_t is None or pcu_t is None:
            continue

        ivr_t = iva_rate(row[idx['IVA']])

        qty = 0
        for v in vcols:
            if v < len(row):
                n = parse_num(row[v])
                if n is not None:
                    qty += int(n)
        if qty <= 0:
            continue

        pvp_x = cx.get('pvp')
        pcu_x = cx.get('pcu')
        if pcu_x is None:
            continue

        # Rentabilidade Real: PVP do TXT (sem IVA) menos o PCU do TXT
        rent_txt_unit = pvp_t / (1 + ivr_t) - pcu_t
        # Rentabilidade Cardex: mesmo PVP real do TXT (sem IVA) menos o PCU do Cardex
        rent_xls_unit = pvp_t / (1 + ivr_t) - pcu_x

        rent_txt_total = rent_txt_unit * qty
        rent_xls_total = rent_xls_unit * qty
        # Diferença positiva => o Cardex é mais rentável (a farmácia ganha mais connosco)
        diff_unit = rent_xls_unit - rent_txt_unit
        diff_total = rent_xls_total - rent_txt_total

        lab = cx.get('laboratorio') or 'Sem Laboratório'
        nome = row[idx['NOM']].strip() if 'NOM' in idx and idx['NOM'] < len(row) else cx.get('descricao')

        products.append({
            'cpr': cpr,
            'nome': nome,
            'laboratorio': lab,
            'qty': qty,
            'pvp_txt': round2(pvp_t),
            'pcu_txt': round2(pcu_t),
            'iva': round(ivr_t, 4),
            'pvp_xls': round2(pvp_x),
            'pcu_xls': round2(pcu_x),
            'rent_txt_unit': round(rent_txt_unit, 4),
            'rent_xls_unit': round(rent_xls_unit, 4),
            'rent_txt_total': round2(rent_txt_total),
            'rent_xls_total': round2(rent_xls_total),
            'diff_unit': round(diff_unit, 4),
            'diff_total': round2(diff_total),
        })

        if lab not in labs_agg:
            labs_agg[lab] = {'laboratorio': lab, 'rent_txt_total': 0.0, 'rent_xls_total': 0.0,
                             'diff_total': 0.0, 'qty': 0, 'n_products': 0}
        a = labs_agg[lab]
        a['rent_txt_total'] += rent_txt_total
        a['rent_xls_total'] += rent_xls_total
        a['diff_total'] += diff_total
        a['qty'] += qty
        a['n_products'] += 1

    labs = []
    for a in labs_agg.values():
        labs.append({
            'laboratorio': a['laboratorio'],
            'rent_txt_total': round2(a['rent_txt_total']),
            'rent_xls_total': round2(a['rent_xls_total']),
            'diff_total': round2(a['diff_total']),
            'qty': a['qty'],
            'n_products': a['n_products'],
        })
    labs.sort(key=lambda x: x['rent_txt_total'], reverse=True)

    summary = {
        'total_rent_txt': round2(sum(p['rent_txt_total'] for p in products)),
        'total_rent_xls': round2(sum(p['rent_xls_total'] for p in products)),
        'total_diff': round2(sum(p['diff_total'] for p in products)),
        'total_qty': sum(p['qty'] for p in products),
        'n_products': len(products),
        'n_labs': len(labs),
    }
    return {'summary': summary, 'labs': labs, 'products': products}


@api_router.post("/analysis/upload")
async def upload_analysis(file: UploadFile = File(...)):
    if not file.filename.lower().endswith('.txt'):
        raise HTTPException(status_code=400, detail="O ficheiro de vendas deve ser um .txt")

    cardex = await db.cardex.find_one({'_id': 'current'})
    if not cardex:
        raise HTTPException(status_code=400, detail="Carregue primeiro o ficheiro Cardex (Definições) antes de analisar vendas.")

    content = await file.read()
    try:
        result = analyse_txt(content, cardex['products'])
    except Exception as e:
        logger.exception("analysis error")
        raise HTTPException(status_code=400, detail=f"Erro ao processar o ficheiro de vendas: {e}")

    if result['summary']['n_products'] == 0:
        raise HTTPException(status_code=400, detail="Nenhum produto com vendas correspondente ao Cardex foi encontrado.")

    analysis_id = str(uuid.uuid4())
    doc = {
        'id': analysis_id,
        'filename': file.filename,
        'created_at': datetime.now(timezone.utc).isoformat(),
        'cardex_filename': cardex.get('filename'),
        **result,
    }
    await db.analyses.insert_one(doc)
    doc.pop('_id', None)
    return doc


@api_router.get("/analysis/latest")
async def latest_analysis():
    doc = await db.analyses.find_one({}, {'_id': 0}, sort=[('created_at', -1)])
    if not doc:
        return {'exists': False}
    return {'exists': True, **doc}


@api_router.get("/analysis")
async def list_analyses():
    docs = await db.analyses.find({}, {'_id': 0, 'products': 0, 'labs': 0}).sort('created_at', -1).to_list(100)
    return docs


@api_router.get("/analysis/{analysis_id}")
async def get_analysis(analysis_id: str):
    doc = await db.analyses.find_one({'id': analysis_id}, {'_id': 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    return doc


@api_router.delete("/analysis/{analysis_id}")
async def delete_analysis(analysis_id: str):
    res = await db.analyses.delete_one({'id': analysis_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    return {'deleted': True}


@api_router.get("/analysis/{analysis_id}/export")
async def export_analysis(analysis_id: str):
    doc = await db.analyses.find_one({'id': analysis_id}, {'_id': 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1A5C46")
    header_font = Font(bold=True, color="FFFFFF")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    s = doc['summary']
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.append(["Indicador", "Valor"])
    ws1.append(["Ficheiro de vendas", doc.get('filename')])
    ws1.append(["Cardex", doc.get('cardex_filename')])
    ws1.append(["Rentabilidade Real (Vendas) €", s['total_rent_txt']])
    ws1.append(["Rentabilidade Cardex (Referência) €", s['total_rent_xls']])
    ws1.append(["Diferença (Cardex - Real) €", s['total_diff']])
    ws1.append(["Quantidade total (12m)", s['total_qty']])
    ws1.append(["Nº de produtos", s['n_products']])
    ws1.append(["Nº de laboratórios", s['n_labs']])
    style_header(ws1, 2)
    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 30

    ws2 = wb.create_sheet("Laboratórios")
    ws2.append(["Laboratório", "Produtos", "Qtd (12m)", "Rent. Real €", "Rent. Cardex €", "Diferença €"])
    for l in doc['labs']:
        ws2.append([l['laboratorio'], l['n_products'], l['qty'], l['rent_txt_total'], l['rent_xls_total'], l['diff_total']])
    style_header(ws2, 6)
    for col, w in zip("ABCDEF", [28, 12, 12, 16, 16, 14]):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Produtos")
    ws3.append(["Código", "Produto", "Laboratório", "Qtd (12m)", "PVP", "PCU",
                "PVP Cardex", "PCU Cardex", "Rent. Real €", "Rent. Cardex €", "Diferença €"])
    for p in sorted(doc['products'], key=lambda x: (x['laboratorio'], -x['rent_txt_total'])):
        ws3.append([p['cpr'], p['nome'], p['laboratorio'], p['qty'], p['pvp_txt'], p['pcu_txt'],
                    p['pvp_xls'], p['pcu_xls'], p['rent_txt_total'], p['rent_xls_total'], p['diff_total']])
    style_header(ws3, 11)
    for col, w in zip("ABCDEFGHIJK", [12, 42, 22, 10, 10, 10, 11, 11, 14, 15, 13]):
        ws3.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "rentabilidade_laboratorios.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api_router.get("/")
async def root():
    return {
        "message": "Pharma Rentabilidade API",
        "version": "2.0",
        "rent_cardex": "PVP_TXT_sem_IVA - PCU_Cardex",
    }


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()